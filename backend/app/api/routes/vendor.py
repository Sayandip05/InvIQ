"""
Vendor API — Excel delivery upload, upload history, and template download.

Vendors can upload Excel files containing delivery data.
The system parses rows, matches item names, and creates inventory transactions.
"""

import logging
from io import BytesIO
from typing import Optional, List, Dict, Any
from fastapi import APIRouter, Depends, UploadFile, File, Request, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.core.dependencies import get_db, get_current_user
from app.core.rate_limiter import limiter
from app.core.exceptions import ValidationError, AuthorizationError
from app.infrastructure.database.models import User, Location
from app.application.vendor_service import VendorService
from app.application.cache_service import cache_invalidate_pattern

logger = logging.getLogger("smart_inventory.vendor")

router = APIRouter(prefix="/vendor", tags=["Vendor"])


def _require_vendor_role(current_user: User) -> None:
    """Ensure user has vendor or admin role."""
    if current_user.role not in {"vendor", "admin"}:
        raise AuthorizationError("Vendor access required")


def _has_location_access(user: User, target_location_id: int) -> bool:
    """
    Safely check if a user has access to target_location_id.
    Handles None, empty list/string, Python list of ints or strings, or raw JSON string.
    """
    raw = user.location_ids
    if not raw:
        return True  # None or empty means no restriction

    if isinstance(raw, str):
        import json
        try:
            raw = json.loads(raw)
        except Exception:
            raw = [raw]

    if isinstance(raw, (list, set, tuple)):
        allowed_ids = set()
        for item in raw:
            try:
                allowed_ids.add(int(item))
            except (ValueError, TypeError):
                pass
        if not allowed_ids:
            return True
        return target_location_id in allowed_ids

    return True


# ── POST /vendor/upload-delivery ───────────────────────────────────────────

@router.post("/upload-delivery")
@limiter.limit("10/minute")
def upload_delivery(
    request: Request,
    location_id: int = Query(..., description="Target location for this delivery"),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Upload an Excel delivery manifest.

    The file should have columns: item_name, quantity_received, delivery_date (optional), notes (optional).
    Each row creates an inventory transaction (received stock).
    """
    _require_vendor_role(current_user)

    # Validate file type
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise ValidationError("Only .xlsx or .xls files are accepted")

    # Validate location exists and belongs to the caller's organization
    if current_user.org_id is None:
        raise AuthorizationError("User is not assigned to an organization")
    location = db.query(Location).filter(Location.id == location_id, Location.org_id == current_user.org_id).first()

    if not location:
        raise AuthorizationError(f"Location {location_id} not found or does not belong to your organization")

    # Check vendor location access safely
    if not _has_location_access(current_user, location_id):
        raise AuthorizationError("You don't have access to this location")

    # Read file
    content = file.file.read()
    if len(content) > 5 * 1024 * 1024:  # 5MB max
        raise ValidationError("File size must be under 5MB")


    service = VendorService(db)
    result = service.parse_and_process_excel(
        file_content=content,
        filename=file.filename,
        location_id=location_id,
        vendor_user_id=current_user.id,
        org_id=current_user.org_id,
    )

    # Invalidate analytics cache so dashboards reflect newly delivered stock
    if result.get("success"):
        cache_invalidate_pattern("analytics:*")

    return result


# ── GET /vendor/my-uploads ─────────────────────────────────────────────────

@router.get("/my-uploads")
def get_my_uploads(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Get upload history for the current vendor."""
    _require_vendor_role(current_user)

    service = VendorService(db)
    uploads = service.get_uploads_for_vendor(current_user.id)

    return {
        "success": True,
        "data": uploads,
        "total": len(uploads),
    }


# ── GET /vendor/template ──────────────────────────────────────────────────

@router.get("/template")
def download_template(
    current_user: User = Depends(get_current_user),
):
    """Download a blank Excel template for vendor deliveries."""
    _require_vendor_role(current_user)

    try:
        import openpyxl
    except ImportError:
        raise ValidationError("openpyxl is not installed on the server")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Delivery Template"

    # Header row
    headers = ["item_name", "quantity_received", "delivery_date", "notes"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)

    # Example row
    ws.cell(row=2, column=1, value="Paracetamol 500mg")
    ws.cell(row=2, column=2, value=100)
    ws.cell(row=2, column=3, value="2026-03-28")
    ws.cell(row=2, column=4, value="Order #12345")

    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 30

    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    wb.close()

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=delivery_template.xlsx"},
    )


# ── GET /vendor/invoices ──────────────────────────────────────────────────

@router.get("/invoices")
def list_invoices(
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=100),
    status: Optional[str] = Query(None, description="Filter by status (ISSUED, PAID, CANCELLED)"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    List delivery invoices.
    Vendors only see their own invoices; Admins and Managers see all organization invoices.
    """
    _require_vendor_role(current_user)

    from app.infrastructure.database.invoice_repo import InvoiceRepository
    invoice_repo = InvoiceRepository(db)

    # Scoped strictly to caller organization
    if current_user.org_id is None:
        raise AuthorizationError("User is not assigned to an organization")

    vendor_filter = current_user.id if current_user.role == "vendor" else None
    invoices, total = invoice_repo.list_invoices(
        org_id=current_user.org_id,
        vendor_user_id=vendor_filter,
        status=status,
        skip=skip,
        limit=limit,
    )

    data = []
    for inv in invoices:
        vendor_name = inv.vendor.full_name if inv.vendor else (inv.vendor.username if inv.vendor else "Vendor")
        data.append({
            "id": inv.id,
            "invoice_number": inv.invoice_number,
            "invoice_date": str(inv.invoice_date),
            "vendor_user_id": inv.vendor_user_id,
            "vendor_name": vendor_name,
            "vendor_upload_id": inv.vendor_upload_id,
            "items_count": len(inv.line_items) if isinstance(inv.line_items, list) else 0,
            "subtotal": inv.subtotal,
            "tax_amount": inv.tax_amount,
            "total_amount": inv.total_amount,
            "status": inv.status,
            "pdf_url": inv.pdf_url,
            "created_at": str(inv.created_at) if inv.created_at else None,
        })

    return {
        "success": True,
        "data": data,
        "total": total,
        "skip": skip,
        "limit": limit,
    }


# ── GET /vendor/invoices/{invoice_id} ─────────────────────────────────────

@router.get("/invoices/{invoice_id}")
def get_invoice_detail(
    invoice_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Get full details of a specific delivery invoice including all line items."""
    _require_vendor_role(current_user)

    from app.infrastructure.database.invoice_repo import InvoiceRepository
    from app.core.exceptions import NotFoundError, AuthorizationError

    invoice_repo = InvoiceRepository(db)
    invoice = invoice_repo.get_by_id(invoice_id)

    if not invoice:
        raise NotFoundError("VendorInvoice", invoice_id)

    # Vendors cannot view other vendors' invoices
    if current_user.role == "vendor" and invoice.vendor_user_id != current_user.id:
        raise AuthorizationError("You do not have permission to view this invoice")

    # Users can only view invoices belonging to their own organization
    if current_user.org_id is None or invoice.org_id != current_user.org_id:
        raise AuthorizationError("Invoice does not belong to your organization")

    vendor_name = invoice.vendor.full_name if invoice.vendor else (invoice.vendor.username if invoice.vendor else "Vendor")

    return {
        "success": True,
        "data": {
            "id": invoice.id,
            "org_id": invoice.org_id,
            "vendor_user_id": invoice.vendor_user_id,
            "vendor_name": vendor_name,
            "vendor_upload_id": invoice.vendor_upload_id,
            "invoice_number": invoice.invoice_number,
            "invoice_date": str(invoice.invoice_date),
            "line_items": invoice.line_items,
            "subtotal": invoice.subtotal,
            "tax_amount": invoice.tax_amount,
            "total_amount": invoice.total_amount,
            "status": invoice.status,
            "pdf_path": invoice.pdf_path,
            "pdf_url": invoice.pdf_url,
            "created_at": str(invoice.created_at) if invoice.created_at else None,
        },
    }


# ── GET /vendor/invoices/{invoice_id}/pdf ─────────────────────────────────

@router.get("/invoices/{invoice_id}/pdf")
def download_invoice_pdf(
    invoice_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Download the generated delivery invoice PDF.
    Streams from Azure Blob Storage or database fallback bytes.
    """
    _require_vendor_role(current_user)

    from app.infrastructure.database.invoice_repo import InvoiceRepository
    from app.infrastructure.storage.azure_blob_storage import get_storage_service
    from app.application.invoice_pdf_service import InvoicePdfService
    from app.core.exceptions import NotFoundError, AuthorizationError

    invoice_repo = InvoiceRepository(db)
    invoice = invoice_repo.get_by_id(invoice_id)

    if not invoice:
        raise NotFoundError("VendorInvoice", invoice_id)

    if current_user.role == "vendor" and invoice.vendor_user_id != current_user.id:
        raise AuthorizationError("You do not have permission to download this invoice")

    # Users can only download invoices belonging to their own organization
    if current_user.org_id is None or invoice.org_id != current_user.org_id:
        raise AuthorizationError("Invoice does not belong to your organization")


    pdf_bytes = None

    # Try downloading from Azure Blob Storage if available
    storage_service = get_storage_service()
    if invoice.pdf_path and storage_service.is_available:
        pdf_bytes = storage_service.download_file(invoice.pdf_path)

    # Fallback to database binary content
    if not pdf_bytes and invoice.pdf_content:
        pdf_bytes = invoice.pdf_content

    # Fallback to dynamic re-render if needed
    if not pdf_bytes:
        vendor = invoice.vendor
        location = invoice.vendor_upload.location if invoice.vendor_upload else None

        vendor_data = {
            "username": vendor.username if vendor else "vendor",
            "full_name": vendor.full_name if vendor else "Vendor",
            "email": vendor.email if vendor else "",
        }
        location_data = {
            "name": location.name if location else "Receiving Location",
            "region": location.region if location else "General",
        }

        invoice_payload = {
            "invoice_number": invoice.invoice_number,
            "invoice_date": invoice.invoice_date,
            "line_items": invoice.line_items,
            "subtotal": invoice.subtotal,
            "tax_amount": invoice.tax_amount,
            "total_amount": invoice.total_amount,
            "status": invoice.status,
        }

        pdf_bytes = InvoicePdfService.generate_invoice_pdf(
            invoice_data=invoice_payload,
            vendor_data=vendor_data,
            location_data=location_data,
        )

    filename = f"{invoice.invoice_number}.pdf"
    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ── GET /vendor/uploads/{upload_id}/invoice ───────────────────────────────

@router.get("/uploads/{upload_id}/invoice")
def get_invoice_by_upload(
    upload_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Retrieve invoice record for a specific vendor upload."""
    _require_vendor_role(current_user)

    from app.infrastructure.database.invoice_repo import InvoiceRepository
    from app.core.exceptions import NotFoundError, AuthorizationError

    invoice_repo = InvoiceRepository(db)
    invoice = invoice_repo.get_by_upload_id(upload_id)

    if not invoice:
        raise NotFoundError("VendorInvoice for upload", upload_id)

    if current_user.role == "vendor" and invoice.vendor_user_id != current_user.id:
        raise AuthorizationError("You do not have permission to view this invoice")

    return {
        "success": True,
        "data": {
            "id": invoice.id,
            "invoice_number": invoice.invoice_number,
            "invoice_date": str(invoice.invoice_date),
            "vendor_user_id": invoice.vendor_user_id,
            "vendor_upload_id": invoice.vendor_upload_id,
            "line_items": invoice.line_items,
            "subtotal": invoice.subtotal,
            "tax_amount": invoice.tax_amount,
            "total_amount": invoice.total_amount,
            "status": invoice.status,
            "pdf_url": invoice.pdf_url,
            "created_at": str(invoice.created_at) if invoice.created_at else None,
        },
    }

